"""
Pipeline Common — shared utilities for URL Consistency Check runners
=====================================================================
Contains:
  - .env loader
  - LLM provider factories (Anthropic, OpenAI, Azure)
  - Web search factory
  - LLM cache load/save
  - Explanation builder
  - Excel formatter
  - Input loader
  - Path constants

Both run_full_check.py and run_flagged_check.py import from here.
"""

import json
import os
import re
import logging
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)


# =============================================================================
#  PATH CONSTANTS
# =============================================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_DIR    = PROJECT_ROOT / "input"
OUTPUT_DIR   = PROJECT_ROOT / "output"
CACHE_DIR    = PROJECT_ROOT / "cache"
CHILD_TABLE_PATH = INPUT_DIR / "issuer_child_flagged.xlsx"
LLM_CACHE_PATH   = CACHE_DIR / "llm_verdicts.json"
ENV_FILE          = PROJECT_ROOT / ".env"

# Rows with CHECK/FLAG==0 are not validated — stamped with this verdict
SKIP_VERDICT = "Don't validate"
SKIP_EXPLANATION_CHECK = (
    "Skipped — URL_CONSISTENCY_CHECK==0 (not flagged for validation)"
)
SKIP_EXPLANATION_FLAG = (
    "Skipped — URL_CONSISTENCY_FLAG==0 (not flagged for validation)"
)


def rows_to_skip_mask(df: pd.DataFrame) -> pd.Series:
    """
    True for rows that should NOT be validated.

    Uses URL_CONSISTENCY_CHECK==0 when present, else numeric URL_CONSISTENCY_FLAG==0.
    """
    if "URL_CONSISTENCY_CHECK" in df.columns:
        return pd.to_numeric(df["URL_CONSISTENCY_CHECK"], errors="coerce").fillna(0).astype(int) == 0
    if "URL_CONSISTENCY_FLAG" in df.columns:
        col = df["URL_CONSISTENCY_FLAG"]
        if pd.api.types.is_numeric_dtype(col):
            return pd.to_numeric(col, errors="coerce").fillna(0).astype(int) == 0
    return pd.Series(False, index=df.index)


def skip_explanation_for_df(df: pd.DataFrame) -> str:
    if "URL_CONSISTENCY_CHECK" in df.columns:
        return SKIP_EXPLANATION_CHECK
    return SKIP_EXPLANATION_FLAG


def stamp_skipped_rows(df: pd.DataFrame, skip_mask: pd.Series = None) -> pd.DataFrame:
    """Mark rows as Don't validate (in-place on copy)."""
    df = df.copy()
    if skip_mask is None:
        skip_mask = rows_to_skip_mask(df)
    if not skip_mask.any():
        return df
    explanation = skip_explanation_for_df(df)
    df.loc[skip_mask, "verdict"] = SKIP_VERDICT
    df.loc[skip_mask, "explanation"] = explanation
    return df


# =============================================================================
#  .ENV LOADER
# =============================================================================

def load_dotenv():
    """Load environment variables from .env file if it exists."""
    if ENV_FILE.exists():
        with open(ENV_FILE) as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    key, val = line.split("=", 1)
                    key, val = key.strip(), val.strip().strip("'\"")
                    os.environ.setdefault(key, val)
        log.info(f"Loaded environment from {ENV_FILE}")


# =============================================================================
#  LLM PROVIDERS
# =============================================================================

def create_anthropic_llm(model: str = "claude-sonnet-4-20250514"):
    """Create an LLM function using Anthropic's Claude API."""
    try:
        import anthropic
    except ImportError:
        raise ImportError("pip install anthropic  (required for --provider anthropic)")

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError(
            "ANTHROPIC_API_KEY not found. Set it via:\n"
            "  export ANTHROPIC_API_KEY=sk-ant-...\n"
            "  or add to .env file in project root"
        )

    client = anthropic.Anthropic(api_key=api_key)

    def llm_fn(prompt: str) -> str:
        response = client.messages.create(
            model=model,
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.content[0].text

    log.info(f"Anthropic LLM ready (model: {model})")
    return llm_fn


def create_openai_llm(model: str = "gpt-4o"):
    """Create an LLM function using OpenAI's API."""
    try:
        import openai
    except ImportError:
        raise ImportError("pip install openai  (required for --provider openai)")

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise ValueError(
            "OPENAI_API_KEY not found. Set it via:\n"
            "  export OPENAI_API_KEY=sk-...\n"
            "  or add to .env file in project root"
        )

    client = openai.OpenAI(api_key=api_key)

    def llm_fn(prompt: str) -> str:
        response = client.chat.completions.create(
            model=model,
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.choices[0].message.content

    log.info(f"OpenAI LLM ready (model: {model})")
    return llm_fn


def create_azure_openai_llm(model: str = None):
    """Create an LLM function using Azure OpenAI Service."""
    try:
        import openai
    except ImportError:
        raise ImportError("pip install openai  (required for --provider azure)")

    api_key = os.environ.get("AZURE_OPENAI_API_KEY")
    endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT")
    deployment = model or os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")
    api_version = os.environ.get("AZURE_OPENAI_API_VERSION", "2024-12-01-preview")

    missing = [k for k, v in {"AZURE_OPENAI_API_KEY": api_key, "AZURE_OPENAI_ENDPOINT": endpoint}.items() if not v]
    if missing:
        raise ValueError(
            f"Missing Azure OpenAI config: {', '.join(missing)}\n"
            "Set them in your .env file in the project root."
        )

    client = openai.AzureOpenAI(
        api_key=api_key,
        azure_endpoint=endpoint,
        api_version=api_version,
    )

    is_reasoning_model = any(deployment.lower().startswith(p) for p in ("o1", "o3"))

    def llm_fn(prompt: str) -> str:
        kwargs = {"model": deployment, "messages": [{"role": "user", "content": prompt}]}
        if is_reasoning_model:
            kwargs["max_completion_tokens"] = 1000
        else:
            kwargs["max_tokens"] = 300
        response = client.chat.completions.create(**kwargs)
        return response.choices[0].message.content

    log.info(f"Azure OpenAI LLM ready (deployment: {deployment}, endpoint: {endpoint})")
    return llm_fn


def create_vertex_llm(model: str = "claude-sonnet-4-6"):
    """Create an LLM function using Anthropic Claude on Google Cloud Vertex AI.

    Requires:
      - pip install 'anthropic[vertex]' google-auth
      - GOOGLE_APPLICATION_CREDENTIALS pointing to a service account JSON key
      - VERTEX_PROJECT_ID  (e.g. proj-dg-dt-datacollqpt001-msci)
      - VERTEX_REGION      (default: global)
    """
    try:
        from anthropic import AnthropicVertex
    except ImportError:
        raise ImportError(
            "pip install 'anthropic[vertex]' google-auth  "
            "(required for --provider vertex)"
        )

    creds_file = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not creds_file or not os.path.isfile(creds_file):
        raise ValueError(
            "GOOGLE_APPLICATION_CREDENTIALS not set or file not found.\n"
            "Set it in your .env file:\n"
            "  GOOGLE_APPLICATION_CREDENTIALS=/path/to/service-account.json"
        )

    project_id = os.environ.get("VERTEX_PROJECT_ID")
    if not project_id:
        raise ValueError(
            "VERTEX_PROJECT_ID not set. Add to .env:\n"
            "  VERTEX_PROJECT_ID=proj-dg-dt-datacollqpt001-msci"
        )

    region = os.environ.get("VERTEX_REGION", "global")

    client = AnthropicVertex(
        project_id=project_id,
        region=region,
    )

    def llm_fn(prompt: str) -> str:
        response = client.messages.create(
            model=model,
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.content[0].text

    log.info(f"Vertex AI LLM ready (model: {model}, project: {project_id}, region: {region})")
    return llm_fn


def create_llm_fn(provider: str, model: str = None):
    """Factory: create the right LLM function based on provider string."""
    if provider == "anthropic":
        return create_anthropic_llm(model or "claude-sonnet-4-20250514")
    elif provider == "vertex":
        return create_vertex_llm(model or "claude-sonnet-4-6")
    elif provider == "azure":
        return create_azure_openai_llm(model=model)
    elif provider == "openai":
        return create_openai_llm(model or "gpt-4o")
    else:
        raise ValueError(f"Unknown LLM provider: {provider}")


# =============================================================================
#  WEB SEARCH
# =============================================================================

def create_search_fn():
    """
    Create a web search function. Tries DuckDuckGo (no API key needed),
    falls back to Google if GOOGLE_API_KEY + GOOGLE_CSE_ID are set.
    """
    try:
        try:
            from ddgs import DDGS
        except ImportError:
            from duckduckgo_search import DDGS

        def search_fn(query: str) -> str:
            try:
                with DDGS() as ddgs:
                    results = list(ddgs.text(query, max_results=3))
                return "\n".join(
                    f"- {r.get('title', '')}: {r.get('body', '')}" for r in results
                )
            except Exception as e:
                return f"(search failed: {e})"

        log.info("Web search: DuckDuckGo (no API key required)")
        return search_fn
    except ImportError:
        pass

    google_key = os.environ.get("GOOGLE_API_KEY")
    google_cse = os.environ.get("GOOGLE_CSE_ID")
    if google_key and google_cse:
        import urllib.request, urllib.parse

        def search_fn(query: str) -> str:
            try:
                params = urllib.parse.urlencode({"key": google_key, "cx": google_cse, "q": query})
                url = f"https://www.googleapis.com/customsearch/v1?{params}"
                req = urllib.request.Request(url)
                with urllib.request.urlopen(req, timeout=10) as resp:
                    data = json.loads(resp.read())
                items = data.get("items", [])[:3]
                return "\n".join(f"- {r.get('title', '')}: {r.get('snippet', '')}" for r in items)
            except Exception as e:
                return f"(search failed: {e})"

        log.info("Web search: Google Custom Search")
        return search_fn

    log.warning(
        "No web search available. Install duckduckgo-search (pip install duckduckgo-search) "
        "or set GOOGLE_API_KEY + GOOGLE_CSE_ID for better LLM verification."
    )
    return None


# =============================================================================
#  CACHE HANDLING
# =============================================================================

def load_llm_cache() -> dict:
    """Load cached LLM verdicts from cache/llm_verdicts.json.
    Returns dict keyed by (ISSUER_ID, domain_stem) → verdict string."""
    if LLM_CACHE_PATH.exists():
        with open(LLM_CACHE_PATH, "r") as f:
            raw = json.load(f)
        verdicts = {}
        for key, val in raw.items():
            parts = key.split("|||")
            if len(parts) == 2:
                verdicts[tuple(parts)] = val
        log.info(f"Loaded {len(verdicts)} cached LLM verdicts from {LLM_CACHE_PATH}")
        return verdicts
    return {}


def save_llm_cache(verdicts: dict):
    """Save LLM verdicts to cache/llm_verdicts.json."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    serializable = {f"{k1}|||{k2}": val for (k1, k2), val in verdicts.items()}
    with open(LLM_CACHE_PATH, "w") as f:
        json.dump(serializable, f, indent=2)
    log.info(f"Saved {len(verdicts)} LLM verdicts to {LLM_CACHE_PATH}")


def load_llm_reasons() -> dict:
    """Load cached LLM reasons from cache/llm_reasons.json."""
    path = CACHE_DIR / "llm_reasons.json"
    if path.exists():
        with open(path, "r") as f:
            return json.load(f)
    return {}


def save_llm_reasons(reasons: dict):
    """Save LLM reasons to cache/llm_reasons.json."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    path = CACHE_DIR / "llm_reasons.json"
    with open(path, "w") as f:
        json.dump(reasons, f, indent=2)


# =============================================================================
#  EXPLANATION MAPPING
# =============================================================================

RULE_EXPLANATIONS = {
    "AUTO_URLC01_REPORT_HOST": "URL is a known financial filing/report hosting site",
    "MANUAL_REVIEW_CHILD_TABLE": "Subsidiary issuer ID found — replace with the correct issuer ID",
    "MANUAL_REVIEW_UPSTREAM_SUGGESTION": "Upstream system suggests a different issuer for this URL — needs verification",
    "AUTO_URLC01_PARENT_URL": "URL domain matches the issuer's own company URL (from child table)",
    "AUTO_URLC01_SUBSIDIARY": "Flagged as subsidiary by upstream system (URL on company domain)",
    "MANUAL_REVIEW_SUBSIDIARY": "Subsidiary flag set but URL is on a third-party domain — verify URL belongs to issuer",
    "AUTO_URLC01_COMPANY_URL": "Domain matches the company's official URL",
    "AUTO_URLC01_MAJORITY_DOMAIN": "Domain matches the majority domain and issuer name (verified)",
    "AUTO_URLC02_GENERIC": "URL is a generic third-party site (Google Maps, LinkedIn, etc.)",
    "AUTO_URLC02_REGISTRY": "Third-party registry/exchange portal/data aggregator (lists company info but not company-owned)",
    "AUTO_URLC02_GENERIC_HOST": "URL is on a generic hosting platform not owned by the company",
    "MAJORITY_NEEDS_LLM_REVIEW": "Majority domain for this issuer but domain name does not match issuer — needs verification",
    "NEEDS_LLM_REVIEW": "Unresolved — needs manual review",
}


def add_explanations(df: pd.DataFrame, llm_reasons: dict = None, verdict_mode: str = "v1") -> pd.DataFrame:
    """Add human-readable explanation column based on auto_decision + LLM reasons.

    Parameters
    ----------
    verdict_mode : str
        Controls explanation text for mode-specific decisions:
        - "v2": MANUAL_REVIEW_SUBSIDIARY → explains as subsidiary not in child table
                 MANUAL_REVIEW_CHILD_TABLE → explains as URLC03 subsidiary match
        - "v4": Report hosts → "Part of allowed regulatory/exchange list"
        - "v5": Allowlisted report hosts / regulatory portals → URLC03 explanation
                 Subsidiaries → detailed subsidiary remarks
                 Third-party → "Third-party: <reason>"
    """
    df = df.copy()

    # Regulatory portal name lookup for report host explanations
    _REGULATORY_PORTAL_NAMES = {
        'tase.co.il': 'TASE (Tel Aviv Stock Exchange)',
        'mayafiles.tase.co.il': 'TASE Maya (Tel Aviv Stock Exchange filings)',
        'cninfo.com.cn': 'CNINFO (China CSRC / Shenzhen Stock Exchange)',
        'static.cninfo.com.cn': 'CNINFO (China CSRC static CDN)',
        'edinet-fsa.go.jp': 'EDINET (Japan financial filing system)',
        'disclosure2dl.edinet-fsa.go.jp': 'EDINET (Japan filing downloads)',
        'dart.fss.or.kr': 'DART (Korea financial filing system)',
        'hkexnews.hk': 'HKEX (Hong Kong Exchange filings)',
        'bseindia.com': 'BSE India (Bombay Stock Exchange)',
        'nseindia.com': 'NSE India (National Stock Exchange)',
        'nsearchives.nseindia.com': 'NSE India Archives',
        'sec.gov': 'SEC EDGAR (US Securities and Exchange Commission)',
        'edgar.sec.gov': 'SEC EDGAR (US SEC direct)',
        'twse.com.tw': 'TWSE (Taiwan Stock Exchange)',
        'doc.twse.com.tw': 'TWSE (Taiwan Stock Exchange documents)',
        'saudiexchange.sa': 'Tadawul (Saudi Exchange)',
        'pse.com.ph': 'PSE (Philippine Stock Exchange)',
        'edge.pse.com.ph': 'PSE Edge (Philippine Stock Exchange)',
        'idx.co.id': 'IDX (Indonesia Stock Exchange)',
        'bursamalaysia.com': 'Bursa Malaysia',
        'sgx.com': 'SGX (Singapore Exchange)',
        'set.or.th': 'SET (Stock Exchange of Thailand)',
        'asx.com.au': 'ASX (Australian Securities Exchange)',
        'nzx.com': 'NZX (New Zealand Exchange)',
        'kap.org.tr': 'KAP (Turkey Public Disclosure Platform)',
        'sse.com.cn': 'SSE (Shanghai Stock Exchange)',
        'jpx.co.jp': 'JPX (Japan Exchange Group)',
        'epa.gov': 'US EPA (Environmental Protection Agency)',
        'fdic.gov': 'US FDIC (Federal Deposit Insurance Corporation)',
        'find-and-update.company-information.service.gov.uk': 'UK Companies House',
        'cloudfront.net': 'AWS CloudFront (IR document hosting CDN)',
        'q4cdn.com': 'Q4 Inc. (IR document hosting)',
        'mziq.com': 'MZ Group (IR platform)',
        'publitas.com': 'Publitas (digital publication CDN)',
        'azurefd.net': 'Azure Front Door (IR document CDN)',
        'irwebpage.com': 'IR Webpage (investor relations platform)',
        'markitdigital.com': 'Markit Digital (ASX research CDN)',
        'annualreports.com': 'AnnualReports.com (report hosting)',
        'listedcompany.com': 'ListedCompany.com (annual report hosting)',
        # ── New report host domains (analyst feedback) ──
        'nasdaq.com': 'NASDAQ (US exchange filings)',
        'euronext.com': 'Euronext (European exchange filings)',
        'live.euronext.com': 'Euronext Live (European exchange filings)',
        'londonstockexchange.com': 'LSE (London Stock Exchange)',
        'www.londonstockexchange.com': 'LSE (London Stock Exchange)',
        'www.rns-pdf.londonstockexchange.com': 'LSE RNS (Regulatory News Service PDFs)',
        'rns-pdf.londonstockexchange.com': 'LSE RNS (Regulatory News Service PDFs)',
        'kind.krx.co.kr': 'KIND (Korea Exchange disclosure)',
        'sedarplus.ca': 'SEDAR+ (Canadian securities filings)',
        'www.sedarplus.ca': 'SEDAR+ (Canadian securities filings)',
        'cnmv.es': 'CNMV (Spanish Securities Commission)',
        'bmv.com.mx': 'BMV (Bolsa Mexicana de Valores)',
        'amf-france.org': 'AMF (French Financial Markets Authority)',
        'senspdf.jse.co.za': 'JSE SENS (South Africa exchange filings)',
        'clientportal.jse.co.za': 'JSE Client Portal (South Africa exchange)',
        'sec.or.th': 'SEC Thailand (Securities and Exchange Commission)',
        'market.sec.or.th': 'SEC Thailand Market Filings',
        'otcmarkets.com': 'OTC Markets Group (US)',
        'www.otcmarkets.com': 'OTC Markets Group (US)',
        'archive.fast-edgar.com': 'FAST-EDGAR (SEC filing archive)',
    }

    def _explain(row):
        decision = row.get("auto_decision", "")

        # ── V4 / V5 mode: custom remarks ──
        if verdict_mode == "v4":
            return _explain_v4(row, decision, _REGULATORY_PORTAL_NAMES, llm_reasons)
        if verdict_mode == "v5":
            return _explain_v5(row, decision, _REGULATORY_PORTAL_NAMES, llm_reasons)

        # Special handling for report host — identify the specific regulatory portal
        if decision == "AUTO_URLC01_REPORT_HOST":
            url = str(row.get("RELEVANT_URL", "")).lower()
            for domain, portal_name in _REGULATORY_PORTAL_NAMES.items():
                if domain in url:
                    return f"Regulatory filing / report host: {portal_name}"
            return RULE_EXPLANATIONS[decision]

        # V2 mode: subsidiary detected but NOT in child table → URLC01
        if verdict_mode == "v2" and decision == "MANUAL_REVIEW_SUBSIDIARY":
            return "Subsidiary detected (upstream flag) — not part of child issuer table; URL treated as valid for parent issuer"

        if verdict_mode == "v2" and decision == "AUTO_URLC01_SUBSIDIARY":
            return "Subsidiary detected (upstream flag) — not part of child issuer table; URL on company domain"

        # Special handling for child table — include matched child issuer details
        if decision == "MANUAL_REVIEW_CHILD_TABLE":
            sug_id = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_ID", "")
            sug_name = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_NAME", "")
            if sug_id and not pd.isna(sug_id):
                parts = [f"Subsidiary issuer ID found: {sug_id}"]
                if sug_name and not pd.isna(sug_name):
                    parts.append(f"({sug_name})")
                if verdict_mode == "v2":
                    parts.append("— subsidiary in child issuer table, replace with correct issuer ID")
                else:
                    parts.append("— replace with the correct issuer ID")
                return " ".join(parts)
            return RULE_EXPLANATIONS[decision]
        # Special handling for upstream suggestion — include suggested issuer details
        if decision == "MANUAL_REVIEW_UPSTREAM_SUGGESTION":
            sug_id = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_ID", "")
            sug_name = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_NAME", "")
            basis = row.get("URL_CONSISTENCY_SUGGESTION_BASIS", "")
            if sug_id and not pd.isna(sug_id):
                parts = [f"Upstream suggests issuer: {sug_id}"]
                if sug_name and not pd.isna(sug_name):
                    parts.append(f"({sug_name})")
                if basis and not pd.isna(basis):
                    parts.append(f"[basis: {basis}]")
                parts.append("— verify and replace issuer ID")
                return " ".join(parts)
            return RULE_EXPLANATIONS[decision]
        if decision in RULE_EXPLANATIONS:
            return RULE_EXPLANATIONS[decision]
        if decision.startswith("LLM_") and llm_reasons:
            # Check for LLM reason in cache
            key = f"{row.get('ISSUER_NAME', '')}|||{row.get('domain_stem', '')}"
            reason = llm_reasons.get(key)
            if not reason:
                key2 = f"{row.get('ISSUER_ID', '')}|||{row.get('domain_stem', '')}"
                reason = llm_reasons.get(key2)
            # Add prefix for majority-domain LLM verified
            if decision.startswith("LLM_MAJ_"):
                prefix = "Majority domain verified by LLM: "
                return prefix + (reason if reason else decision.replace("LLM_MAJ_", ""))
            return reason if reason else f"LLM verified ({decision})"
        if decision == "MAJORITY_NEEDS_LLM_REVIEW":
            return RULE_EXPLANATIONS[decision]
        return "Unresolved — needs manual review"

    df["explanation"] = df.apply(_explain, axis=1)
    return df


def _has_upstream_domain_mismatch(row) -> bool:
    """True when upstream flags indicate the URL domain does not match the issuer."""
    for col in ("URL_CONSISTENCY_FLAG", "URL_CONSISTENCY_DETAILS"):
        val = row.get(col, "")
        if val is not None and not pd.isna(val) and "DOMAIN_MISMATCH" in str(val):
            return True
    return False


_COMPANY_URL_COL_PATTERNS = (
    "COMPANY_URL", "PARENT_URL", "ISSUER_URL", "OFFICIAL_URL",
    "OFFICIAL_WEBSITE", "WEBSITE", "CORPORATE_URL",
)
_COMPANY_DOC_PATH_RE = re.compile(
    r"annual[-_]?report|financial[-_]?report|/ir/|/investor|"
    r"investor[-_]?relation|/ar/|/reports?/|10-k|10-q|prospectus|\.pdf(?:\?|$)",
    re.I,
)


def _v4_is_company_owned(row) -> bool:
    """
    True when RELEVANT_URL is on the issuer's own domain (not a third-party CDN/portal).

    Covers: parent URL match, company URL column, majority-verified domain,
    subsidiary on company domain (even if not in child table), and company-hosted PDFs.
    """
    from url_consistency_engine import (
        extract_root_domain,
        extract_domain_stem,
        _is_third_party_registry,
        _is_generic_url,
    )

    url = row.get("RELEVANT_URL", "")
    if pd.isna(url) or not str(url).strip():
        return False
    url = str(url)
    if _is_third_party_registry(url) or _is_generic_url(url):
        return False

    decision = row.get("auto_decision", "")
    if decision in (
        "AUTO_URLC01_PARENT_URL",
        "AUTO_URLC01_COMPANY_URL",
        "AUTO_URLC01_MAJORITY_DOMAIN",
        "AUTO_URLC01_SUBSIDIARY",
    ):
        return True

    url_root = extract_root_domain(url)
    url_stem = row.get("domain_stem") or extract_domain_stem(url)

    for col in row.index:
        col_upper = str(col).upper()
        if not any(p in col_upper for p in _COMPANY_URL_COL_PATTERNS):
            if col_upper != "URL":
                continue
        co = row.get(col)
        if co is None or pd.isna(co) or not str(co).strip():
            continue
        co_str = str(co).strip()
        if not co_str.startswith(("http://", "https://")):
            co_str = "https://" + co_str
        co_root = extract_root_domain(co_str)
        co_stem = extract_domain_stem(co_str)
        if co_root and url_root and (url_root == co_root or url_root.endswith("." + co_root)):
            return True
        if co_stem and url_stem and co_stem == url_stem:
            return True

    return False


def _v4_is_company_document(row) -> bool:
    """True when URL path looks like a company annual report, filing, or PDF."""
    url = row.get("RELEVANT_URL", "")
    if pd.isna(url) or not str(url).strip():
        return False
    return bool(_COMPANY_DOC_PATH_RE.search(str(url)))


def _v5_is_allowlisted_third_party(row) -> bool:
    """
    True when RELEVANT_URL is on the regulatory/exchange/government allowlist
    (REPORT_HOST_DOMAINS) — acceptable third-party filing portals, not company-owned.
    """
    from url_consistency_engine import _is_report_host

    url = row.get("RELEVANT_URL", "")
    if pd.isna(url) or not str(url).strip():
        return False
    return _is_report_host(str(url))


def _allowlist_portal_explanation(row, portal_names: dict) -> str:
    """Human-readable explanation for an allowlisted regulatory/exchange/gov portal."""
    url = str(row.get("RELEVANT_URL", "")).lower()
    for domain, portal_name in portal_names.items():
        if domain in url:
            return f"Allowed third-party (URLC03): regulatory/exchange/government portal — {portal_name}"
    return "Allowed third-party (URLC03): regulatory/exchange/government portal on allowlist"


def _explain_v4(row, decision, portal_names, llm_reasons):
    """V4-specific explanation builder with detailed remarks."""

    # ── Report host / regulatory portal → URLC02 with allowed-list remark ──
    if decision == "AUTO_URLC01_REPORT_HOST":
        url = str(row.get("RELEVANT_URL", "")).lower()
        for domain, portal_name in portal_names.items():
            if domain in url:
                return f"Third-party: Part of allowed regulatory/exchange list — {portal_name}"
        return "Third-party: Part of allowed regulatory/exchange list"

    # ── Generic third-party ──
    if decision == "AUTO_URLC02_GENERIC":
        return "Third-party: Generic site (Google Maps, LinkedIn, social media, etc.) — not on allowed list"

    # ── Registry / aggregator ──
    if decision == "AUTO_URLC02_REGISTRY":
        return "Third-party: Registry/exchange portal/data aggregator — not on allowed list"

    # ── Generic hosting ──
    if decision == "AUTO_URLC02_GENERIC_HOST":
        return "Third-party: Generic hosting platform — not on allowed list"

    # ── Subsidiary found in child table ──
    if decision == "MANUAL_REVIEW_CHILD_TABLE":
        sug_id = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_ID", "")
        sug_name = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_NAME", "")
        parts = ["Subsidiary URL"]
        if sug_id and not pd.isna(sug_id):
            parts.append(f"— found in child issuer table as {sug_id}")
            if sug_name and not pd.isna(sug_name):
                parts.append(f"({sug_name})")
            parts.append("— needs reassignment to correct issuer ID")
        else:
            parts.append("— found in child issuer table but no issuer ID available")
        return " ".join(parts)

    # ── Subsidiary flag (not in child table, NOT on company domain) ──
    if decision == "MANUAL_REVIEW_SUBSIDIARY":
        return "Subsidiary URL — upstream flag set but subsidiary NOT found in child issuer table; URL not on company domain — needs verification"

    # ── Subsidiary flag + company domain ──
    if decision == "AUTO_URLC01_SUBSIDIARY":
        return "Subsidiary URL — upstream flag set but subsidiary NOT found in child issuer table; URL is on company domain — verify subsidiary relationship"

    # ── Upstream suggestion ──
    if decision == "MANUAL_REVIEW_UPSTREAM_SUGGESTION":
        sug_id = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_ID", "")
        sug_name = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_NAME", "")
        basis = row.get("URL_CONSISTENCY_SUGGESTION_BASIS", "")
        parts = ["Upstream system suggests different issuer"]
        if sug_id and not pd.isna(sug_id):
            parts.append(f"— suggested: {sug_id}")
            if sug_name and not pd.isna(sug_name):
                parts.append(f"({sug_name})")
        if basis and not pd.isna(basis):
            parts.append(f"[basis: {basis}]")
        parts.append("— verify and reassign")
        return " ".join(parts)

    # ── Direct company matches (URLC01) ──
    if decision == "AUTO_URLC01_PARENT_URL":
        return "Company-owned: Domain matches the issuer's own company URL (from child table)"
    if decision == "AUTO_URLC01_COMPANY_URL":
        return "Company-owned: Domain matches the company's official URL"
    if decision == "AUTO_URLC01_MAJORITY_DOMAIN":
        return "Company-owned: Domain matches the majority domain and issuer name (verified)"

    # ── LLM verdicts ──
    if decision.startswith("LLM_") and llm_reasons:
        key = f"{row.get('ISSUER_NAME', '')}|||{row.get('domain_stem', '')}"
        reason = llm_reasons.get(key)
        if not reason:
            key2 = f"{row.get('ISSUER_ID', '')}|||{row.get('domain_stem', '')}"
            reason = llm_reasons.get(key2)

        if decision in ("LLM_URLC02", "LLM_MAJ_URLC02"):
            prefix = "Third-party (LLM-detected): "
            return prefix + (reason if reason else "not on allowed list")
        if decision in ("LLM_URLC01", "LLM_MAJ_URLC01"):
            if _v4_is_company_owned(row):
                prefix = "Company-owned (LLM-verified): "
                if _v4_is_company_document(row):
                    return prefix + (reason if reason else "company annual report or filing on own domain")
                return prefix + (reason if reason else "direct company association on own domain")
            prefix = "Third-party (v4 — not company-owned domain): "
            return prefix + (reason if reason else "hosted on third-party domain, not issuer's own site")

    if decision == "MAJORITY_NEEDS_LLM_REVIEW":
        return "Majority domain but name mismatch — needs verification"

    return "Unresolved — needs manual review"


def _explain_v5(row, decision, portal_names, llm_reasons):
    """V5-specific explanation builder — same as v4 but allowlisted third-party → URLC03."""

    if decision == "AUTO_URLC01_REPORT_HOST":
        return _allowlist_portal_explanation(row, portal_names)

    if decision in ("AUTO_URLC02_GENERIC", "AUTO_URLC02_REGISTRY", "AUTO_URLC02_GENERIC_HOST"):
        if _v5_is_allowlisted_third_party(row):
            return _allowlist_portal_explanation(row, portal_names)
        if decision == "AUTO_URLC02_GENERIC":
            return "Third-party: Generic site (Google Maps, LinkedIn, social media, etc.) — not on allowlist"
        if decision == "AUTO_URLC02_REGISTRY":
            return "Third-party: Registry/exchange portal/data aggregator — not on allowlist"
        return "Third-party: Generic hosting platform — not on allowlist"

    # Subsidiary / upstream / company-owned explanations — same as v4
    if decision == "MANUAL_REVIEW_CHILD_TABLE":
        sug_id = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_ID", "")
        sug_name = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_NAME", "")
        parts = ["Subsidiary URL"]
        if sug_id and not pd.isna(sug_id):
            parts.append(f"— found in child issuer table as {sug_id}")
            if sug_name and not pd.isna(sug_name):
                parts.append(f"({sug_name})")
            parts.append("— needs reassignment to correct issuer ID")
        else:
            parts.append("— found in child issuer table but no issuer ID available")
        return " ".join(parts)

    if decision == "MANUAL_REVIEW_SUBSIDIARY":
        return "Subsidiary URL — upstream flag set but subsidiary NOT found in child issuer table; URL not on company domain — needs verification"

    if decision == "AUTO_URLC01_SUBSIDIARY":
        return "Subsidiary URL — upstream flag set but subsidiary NOT found in child issuer table; URL is on company domain — verify subsidiary relationship"

    if decision == "MANUAL_REVIEW_UPSTREAM_SUGGESTION":
        sug_id = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_ID", "")
        sug_name = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_NAME", "")
        basis = row.get("URL_CONSISTENCY_SUGGESTION_BASIS", "")
        parts = ["Upstream system suggests different issuer"]
        if sug_id and not pd.isna(sug_id):
            parts.append(f"— suggested: {sug_id}")
            if sug_name and not pd.isna(sug_name):
                parts.append(f"({sug_name})")
        if basis and not pd.isna(basis):
            parts.append(f"[basis: {basis}]")
        parts.append("— verify and reassign")
        return " ".join(parts)

    if decision == "AUTO_URLC01_PARENT_URL":
        return "Company-owned: Domain matches the issuer's own company URL (from child table)"
    if decision == "AUTO_URLC01_COMPANY_URL":
        return "Company-owned: Domain matches the company's official URL"
    if decision == "AUTO_URLC01_MAJORITY_DOMAIN":
        return "Company-owned: Domain matches the majority domain and issuer name (verified)"

    if decision.startswith("LLM_") and llm_reasons:
        key = f"{row.get('ISSUER_NAME', '')}|||{row.get('domain_stem', '')}"
        reason = llm_reasons.get(key)
        if not reason:
            key2 = f"{row.get('ISSUER_ID', '')}|||{row.get('domain_stem', '')}"
            reason = llm_reasons.get(key2)

        if decision in ("LLM_URLC02", "LLM_MAJ_URLC02"):
            if _v5_is_allowlisted_third_party(row):
                prefix = "Allowed third-party (URLC03, LLM-verified): "
            else:
                prefix = "Third-party (LLM-detected): "
            return prefix + (reason if reason else "not on allowlist")
        if decision in ("LLM_URLC01", "LLM_MAJ_URLC01"):
            if _v4_is_company_owned(row):
                prefix = "Company-owned (LLM-verified): "
                if _v4_is_company_document(row):
                    return prefix + (reason if reason else "company annual report or filing on own domain")
                return prefix + (reason if reason else "direct company association on own domain")
            if _v5_is_allowlisted_third_party(row):
                prefix = "Allowed third-party (URLC03, LLM-verified): "
                return prefix + (reason if reason else "regulatory/exchange/government portal on allowlist")
            prefix = "Third-party (not company-owned domain): "
            return prefix + (reason if reason else "hosted on third-party domain, not on allowlist")

    if decision == "MAJORITY_NEEDS_LLM_REVIEW":
        return "Majority domain but name mismatch — needs verification"

    return "Unresolved — needs manual review"


# =============================================================================
#  V4 TAG COLUMN — Corporate Relationship Classification
# =============================================================================

def build_reverse_parent_lookup(child_df: pd.DataFrame) -> dict:
    """
    Build a reverse lookup: child_issuer_id → list of parent issuer_ids.

    Used to determine if the current issuer is itself a subsidiary (child)
    of another listed entity, enabling 'ultimate_parent' / 'parent' tag assignment.

    Returns
    -------
    dict : child_issuer_id → [{'parent_id': ..., 'parent_name': ..., 'parent_url': ...}, ...]
    """
    from url_consistency_engine import extract_domain_stem

    reverse = {}
    for _, row in child_df.iterrows():
        cid = row.get('child_issuer_id')
        if pd.isna(cid):
            continue
        cid = str(cid)
        entry = {
            'parent_id': str(row.get('issuer_id', '')),
            'parent_name': str(row.get('name', '')),
            'parent_url': str(row.get('url', '')),
            'parent_stem': extract_domain_stem(str(row.get('url', ''))),
        }
        if cid not in reverse:
            reverse[cid] = []
        # Avoid duplicate parents
        if not any(e['parent_id'] == entry['parent_id'] for e in reverse[cid]):
            reverse[cid].append(entry)
    return reverse


def build_sibling_lookup(child_df: pd.DataFrame) -> dict:
    """
    Build a lookup: issuer_id → set of all child_issuer_ids (siblings in the same family).

    Used to detect 'affiliate_not_subsidiary' — when the URL belongs to a sibling
    entity under the same parent, not in the direct parent→child lineage.
    """
    from url_consistency_engine import extract_domain_stem

    # Map: parent_issuer_id → list of (child_issuer_id, child_stem)
    family = {}
    for _, row in child_df.iterrows():
        pid = str(row.get('issuer_id', ''))
        cid = row.get('child_issuer_id')
        if pd.isna(cid):
            continue
        curl = str(row.get('child_url', ''))
        cstem = extract_domain_stem(curl)
        if pid not in family:
            family[pid] = []
        family[pid].append({'child_id': str(cid), 'child_stem': cstem})
    return family


def add_v4_tags(result_df: pd.DataFrame, child_df: pd.DataFrame) -> pd.DataFrame:
    """
    Add the 'tag' column for v4 verdict mode.

    Tags describe the corporate relationship between the URL's owner entity
    and the issuer being checked:

      real_subsidiary_with_issuer_id   — URL belongs to a confirmed subsidiary
                                         that has a known issuer_id in the child table
      real_subsidiary_without_issuer_id — URL belongs to a confirmed subsidiary,
                                         but that subsidiary has no issuer_id in our database
      ultimate_parent                  — The URL belongs to the ultimate parent of
                                         this issuer (issuer is itself a subsidiary)
      parent                           — The URL belongs to an ancestor (not the ultimate)
      affiliate_not_subsidiary         — Same corporate family but not in direct lineage
      unrelated_entity                 — Completely different corporate family (third-party)
    """
    from url_consistency_engine import extract_domain_stem

    df = result_df.copy()

    # Build lookups
    log.info("Building v4 tag lookups...")
    reverse_parent = build_reverse_parent_lookup(child_df)
    sibling_families = build_sibling_lookup(child_df)

    # Pre-compute: for each issuer, gather all sibling stems (children of their parent)
    # issuer_id → set of sibling domain stems
    _issuer_sibling_stems = {}
    for iid, parents in reverse_parent.items():
        sibling_stems = set()
        for p in parents:
            pid = p['parent_id']
            if pid in sibling_families:
                for sib in sibling_families[pid]:
                    if sib['child_id'] != iid and sib['child_stem']:
                        sibling_stems.add(sib['child_stem'])
        if sibling_stems:
            _issuer_sibling_stems[iid] = sibling_stems

    def _assign_tag(row):
        decision = row.get("auto_decision", "")
        issuer_id = str(row.get("ISSUER_ID", ""))
        domain_stem = row.get("domain_stem", "")

        # ── Subsidiary found in child table ──
        if decision == "MANUAL_REVIEW_CHILD_TABLE":
            sug_id = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_ID", "")
            if sug_id and not pd.isna(sug_id) and str(sug_id).startswith("IID"):
                return "real_subsidiary_with_issuer_id"
            else:
                return "real_subsidiary_without_issuer_id"

        # ── Subsidiary flag set (not in child table) ──
        if decision in ("MANUAL_REVIEW_SUBSIDIARY", "AUTO_URLC01_SUBSIDIARY"):
            return "real_subsidiary_without_issuer_id"

        # ── Third-party URLs → unrelated_entity ──
        if decision in ("AUTO_URLC02_GENERIC", "AUTO_URLC02_REGISTRY",
                         "AUTO_URLC02_GENERIC_HOST"):
            return "unrelated_entity"

        # ── Report host → unrelated_entity (third-party portal in v4) ──
        if decision == "AUTO_URLC01_REPORT_HOST":
            return "unrelated_entity"

        # ── LLM-detected third-party → unrelated_entity ──
        if decision in ("LLM_URLC02", "LLM_MAJ_URLC02"):
            return "unrelated_entity"

        # ── Direct company match → check if it's actually a parent relationship ──
        # If our issuer is itself a child in the child table, and the URL
        # belongs to its parent company, tag as ultimate_parent or parent
        if decision in ("AUTO_URLC01_PARENT_URL", "AUTO_URLC01_COMPANY_URL",
                         "AUTO_URLC01_MAJORITY_DOMAIN",
                         "LLM_URLC01", "LLM_MAJ_URLC01"):
            # Check if issuer is someone else's child
            if issuer_id in reverse_parent and domain_stem:
                parents = reverse_parent[issuer_id]
                for p in parents:
                    if p['parent_stem'] and p['parent_stem'] == domain_stem:
                        # URL matches a parent's domain
                        # Check if this parent is the ultimate parent
                        # (ultimate = parent is NOT itself a child of someone else)
                        if p['parent_id'] not in reverse_parent:
                            return "ultimate_parent"
                        else:
                            return "parent"

            # Check if URL belongs to a sibling (affiliate)
            if issuer_id in _issuer_sibling_stems and domain_stem:
                if domain_stem in _issuer_sibling_stems[issuer_id]:
                    return "affiliate_not_subsidiary"

            # Direct company match — no special relationship
            return None  # No tag needed for direct company URLs

        # ── Upstream suggestion → could be subsidiary or parent ──
        if decision == "MANUAL_REVIEW_UPSTREAM_SUGGESTION":
            sug_id = row.get("URL_CONSISTENCY_SUGGESTED_ISSUER_ID", "")
            if sug_id and not pd.isna(sug_id) and str(sug_id).startswith("IID"):
                # Check if suggested issuer is a child of current issuer
                # (meaning URL belongs to a subsidiary)
                if issuer_id in sibling_families:
                    child_ids = {c['child_id'] for c in sibling_families[issuer_id]}
                    if str(sug_id) in child_ids:
                        return "real_subsidiary_with_issuer_id"
                # Check if suggested issuer is a parent
                if issuer_id in reverse_parent:
                    parent_ids = {p['parent_id'] for p in reverse_parent[issuer_id]}
                    if str(sug_id) in parent_ids:
                        if str(sug_id) not in reverse_parent:
                            return "ultimate_parent"
                        return "parent"
                return "real_subsidiary_with_issuer_id"
            return "real_subsidiary_without_issuer_id"

        # ── Fallback ──
        return "unrelated_entity"

    log.info("Assigning v4 tags...")
    df["tag"] = df.apply(_assign_tag, axis=1)
    log.info(f"Tag distribution:\n{df['tag'].value_counts()}")
    return df


# =============================================================================
#  INPUT LOADING
# =============================================================================

def load_input(path: str) -> pd.DataFrame:
    """Load input data from .pkl, .xlsx, or .csv."""
    path = Path(path)
    if path.suffix == ".pkl":
        return pd.read_pickle(path)
    elif path.suffix in (".xlsx", ".xls"):
        return pd.read_excel(path)
    elif path.suffix == ".csv":
        return pd.read_csv(path)
    else:
        raise ValueError(f"Unsupported file format: {path.suffix}")


def resolve_path(user_path: str, default_dir: Path = None) -> Path:
    """Resolve a user-supplied path, making it absolute relative to PROJECT_ROOT."""
    p = Path(user_path)
    if p.is_absolute():
        return p
    return (default_dir or PROJECT_ROOT) / p


def detect_company_url_col(df: pd.DataFrame, col_name: str = None) -> str:
    """Validate and return the company URL column name, or None if not found."""
    if not col_name:
        return None
    if col_name not in df.columns:
        log.warning(f"Company URL column '{col_name}' not found in input. Skipping company URL match.")
        return None
    log.info(f"Using company URL column: {col_name}")
    return col_name


# =============================================================================
#  EXCEL FORMATTING
# =============================================================================

def format_excel(path: str):
    """Apply professional formatting to output Excel file."""
    wb = load_workbook(path)
    ws = wb.active

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    urlc01_font = Font(name="Calibri", color="006100", bold=True)
    urlc01_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    urlc02_font = Font(name="Calibri", color="9C0006", bold=True)
    urlc02_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    urlc03_font = Font(name="Calibri", color="00336B", bold=True)
    urlc03_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    manual_font = Font(name="Calibri", color="9C6500", bold=True)
    manual_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    cols = {cell.value: cell.column for cell in ws[1]}
    verdict_col = cols.get("verdict")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
        if verdict_col:
            v = row[verdict_col - 1]
            val = str(v.value).strip() if v.value else ""
            if val == "URLC01":
                v.font, v.fill = urlc01_font, urlc01_fill
            elif "URLC03" in val:
                v.font, v.fill = urlc03_font, urlc03_fill
            elif "URLC02" in val:
                v.font, v.fill = urlc02_font, urlc02_fill
            elif "Manual Review" in val or "Manual_Review" in val:
                v.font, v.fill = manual_font, manual_fill
            elif val == SKIP_VERDICT or "Don't validate" in val:
                v.font = Font(name="Calibri", color="808080", italic=True)
                v.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    if verdict_col:
        ws.column_dimensions[get_column_letter(verdict_col)].width = 22
    exp_col = cols.get("explanation")
    if exp_col:
        ws.column_dimensions[get_column_letter(exp_col)].width = 60

    # Tag column formatting (v4)
    tag_col = cols.get("tag")
    if tag_col:
        # Color-code tags
        tag_colors = {
            "real_subsidiary_with_issuer_id":    (Font(name="Calibri", color="00336B", bold=True),
                                                   PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")),
            "real_subsidiary_without_issuer_id": (Font(name="Calibri", color="9C6500", bold=True),
                                                   PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")),
            "ultimate_parent":                   (Font(name="Calibri", color="4A148C", bold=True),
                                                   PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")),
            "parent":                            (Font(name="Calibri", color="4A148C", bold=False),
                                                   PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")),
            "affiliate_not_subsidiary":          (Font(name="Calibri", color="E65100", bold=True),
                                                   PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")),
            "unrelated_entity":                  (Font(name="Calibri", color="9C0006", bold=True),
                                                   PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
        }
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            t = row[tag_col - 1]
            tag_val = str(t.value).strip() if t.value else ""
            if tag_val in tag_colors:
                t.font, t.fill = tag_colors[tag_val]
        ws.column_dimensions[get_column_letter(tag_col)].width = 38

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    log.info(f"Formatting applied to {path}")


# =============================================================================
#  CORE PIPELINE RUNNER (shared logic)
# =============================================================================

def run_classification(
    fail_df: pd.DataFrame,
    pass_df: pd.DataFrame = None,
    child_path: str = None,
    company_url_col: str = None,
    skip_llm: bool = True,
    provider: str = None,
    model: str = None,
    output_path: str = None,
    run_eval: bool = False,
    verdict_mode: str = "v1",
):
    """
    Core classification logic used by both full-check and flagged-check runners.

    Parameters
    ----------
    fail_df : DataFrame
        Rows to classify (either all rows or only flagged rows).
    pass_df : DataFrame or None
        Rows that passed the check (skipped). If provided, verdicts are merged
        back into the full dataset with pass_df rows having no verdict.
    child_path : str
        Path to issuer_child_flagged.xlsx.
    company_url_col : str or None
        Column containing the company's official URL.
    skip_llm : bool
        If True, only use cache (no live LLM calls).
    provider : str or None
        LLM provider name ('anthropic', 'openai', 'azure').
    model : str or None
        LLM model override.
    output_path : str
        Where to save the formatted Excel output.
    run_eval : bool
        Whether to run accuracy evaluation (needs COMMENT_CODE column).
    verdict_mode : str
        Verdict output mode — controls which verdict codes appear:
          "v1" — URLC01 & Manual Review only (URLC02 merged into Manual Review)
          "v2" — URLC01, URLC02 (third party), URLC03 (subsidiary in child table),
                  Manual Review (unclear). Subsidiary detected but NOT in child table
                  → URLC01 with explanation.
          "v3" — For URL_CONSISTENCY_FLAG==1 rows: URLC01 & Manual Review only
          "v4" — Strict: URLC01 (direct company only), URLC02 (ALL third-party
                  incl. stock exchanges & gov registries), Manual_Review (ALL
                  subsidiaries). Adds Tag column with corporate relationship.
          "v5" — Like v4, but allowlisted regulatory/exchange/government portals
                  and IR filing hosts → URLC03; other third-party → URLC02.
                  Adds Tag column with corporate relationship.

    Returns
    -------
    dict with summary stats.
    """
    from url_consistency_engine import (
        run_pipeline,
        get_llm_review_combos,
        batch_verify_combos,
        verdicts_to_dict,
        evaluate_against_ground_truth,
    )

    child_path = child_path or str(CHILD_TABLE_PATH)

    # ── All rows skipped (CHECK/FLAG==0) — output without running pipeline ───
    if fail_df.empty:
        if pass_df is None or pass_df.empty:
            log.info("No rows to classify.")
            return {"total": 0, "classified": 0}
        out = stamp_skipped_rows(pass_df.copy())
        output_path = output_path or str(OUTPUT_DIR / "results.xlsx")
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        out.to_excel(str(output_path), index=False)
        log.info(f"Results saved to {output_path} ({len(out)} rows, all '{SKIP_VERDICT}')")
        format_excel(str(output_path))
        return {"total": len(out), "classified": 0, "skipped": len(out)}

    # ── Load child table ─────────────────────────────────────────────────────
    log.info(f"Loading child table from {child_path}")
    child_df = pd.read_excel(str(child_path))
    log.info(f"Loaded {len(child_df)} child records")

    # ── Load caches ──────────────────────────────────────────────────────────
    llm_verdicts = load_llm_cache()
    llm_reasons = load_llm_reasons()

    # ── Validate company URL column ──────────────────────────────────────────
    company_url_col = detect_company_url_col(fail_df, company_url_col)

    # ── Run pipeline (Layer 1+2 + cached LLM verdicts) ───────────────────────
    result = run_pipeline(fail_df, child_df,
                          llm_verdicts=llm_verdicts if llm_verdicts else None,
                          company_url_col=company_url_col)

    # ── Handle unresolved combos ─────────────────────────────────────────────
    combos = get_llm_review_combos(result)

    if not combos.empty and not skip_llm and provider:
        log.info(f"{len(combos)} new combos need LLM verification via {provider}")
        llm_fn = create_llm_fn(provider, model)
        search_fn = create_search_fn()
        results_list = batch_verify_combos(combos, search_fn=search_fn, llm_fn=llm_fn)

        new_verdicts = verdicts_to_dict(results_list)
        llm_verdicts.update(new_verdicts)
        save_llm_cache(llm_verdicts)

        for r in results_list:
            if r.get("reason"):
                key = f"{r['issuer_name']}|||{r['domain_stem']}"
                llm_reasons[key] = r["reason"]
        save_llm_reasons(llm_reasons)

        log.info("Re-running pipeline with new LLM verdicts...")
        result = run_pipeline(fail_df, child_df, llm_verdicts=llm_verdicts,
                              company_url_col=company_url_col)

    elif not combos.empty and not skip_llm and not provider:
        log.warning(
            f"\n{'='*60}\n"
            f"  {len(combos)} combos still need LLM verification.\n"
            f"  Use --provider anthropic|vertex|openai|azure  or add to cache manually.\n"
            f"{'='*60}"
        )
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        combos.to_excel(str(OUTPUT_DIR / "unresolved_combos.xlsx"), index=False)

    # ── Build verdict + explanation columns ───────────────────────────────────
    log.info(f"Applying verdict mode: {verdict_mode}")

    if verdict_mode == "v2":
        # ── V2: Granular verdicts ──
        # URLC01 = valid, URLC02 = third party, URLC03 = subsidiary in child table,
        # Manual Review = unclear OR LLM-detected (not in dictionary).
        # Dictionary-based detections (AUTO_URLC01_*) → trusted → URLC01
        # LLM-detected verdicts (LLM_URLC01, LLM_MAJ_URLC01) → Manual Review
        # Special: subsidiary detected but NOT in child table → URLC01
        def _v2_verdict(row):
            decision = row.get("auto_decision", "")
            code = row.get("predicted_url_code", "")
            # Subsidiary found in child table → URLC03
            if decision == "MANUAL_REVIEW_CHILD_TABLE":
                return "URLC03"
            # Subsidiary flag set, URL not on third-party, but not in child table → URLC01
            if decision == "MANUAL_REVIEW_SUBSIDIARY":
                return "URLC01"
            # LLM-detected verdicts → Manual Review (not in hardcoded dictionary)
            if decision in ("LLM_URLC01", "LLM_MAJ_URLC01"):
                return "Manual Review"
            # LLM-detected third-party stays URLC02 (still third-party)
            if decision == "LLM_URLC02":
                return "URLC02"
            # Dictionary-based deterministic rules → trusted
            if code == "URLC01":
                return "URLC01"
            if code == "URLC02":
                return "URLC02"
            if code == "MANUAL_REVIEW":
                return "Manual Review"
            return "Manual Review"  # UNRESOLVED → Manual Review

        result["verdict"] = result.apply(_v2_verdict, axis=1)

    elif verdict_mode == "v3":
        # ── V3: Flag==1 mode — URLC01 & Manual Review only ──
        result["verdict"] = result["predicted_url_code"].replace({
            "UNRESOLVED": "Manual Review",
            "MANUAL_REVIEW": "Manual Review",
            "URLC02": "Manual Review",
        })

    elif verdict_mode == "v4":
        # ── V4: Strict company-only mode ──
        # URLC01  = company's own domain, company-hosted PDFs/annual reports
        # URLC02  = ALL third-party URLs (CDNs, registries, exchanges, aggregators)
        # Manual_Review = subsidiaries (in or not in child table), upstream suggestions
        def _v4_verdict(row):
            decision = row.get("auto_decision", "")
            code = row.get("predicted_url_code", "")

            # ── Third-party (always URLC02) ──
            if decision in ("AUTO_URLC02_GENERIC", "AUTO_URLC02_REGISTRY", "AUTO_URLC02_GENERIC_HOST"):
                return "URLC02"

            # ── Report hosts / regulatory portals → URLC02 in v4 ──
            if decision == "AUTO_URLC01_REPORT_HOST":
                return "URLC02"

            # ── Subsidiary in child table → Manual_Review (reassign to correct issuer) ──
            if decision == "MANUAL_REVIEW_CHILD_TABLE":
                return "Manual_Review"

            # ── Subsidiary flag, not in child table, NOT on company domain ──
            if decision == "MANUAL_REVIEW_SUBSIDIARY":
                return "Manual_Review"

            # ── Subsidiary on company domain (not in child list) → Manual_Review ──
            if decision == "AUTO_URLC01_SUBSIDIARY":
                return "Manual_Review"

            # ── Upstream suggestion → Manual_Review ──
            if decision == "MANUAL_REVIEW_UPSTREAM_SUGGESTION":
                return "Manual_Review"

            # ── Direct company URL match → URLC01 ──
            if decision in (
                "AUTO_URLC01_PARENT_URL",
                "AUTO_URLC01_COMPANY_URL",
                "AUTO_URLC01_MAJORITY_DOMAIN",
            ):
                return "URLC01"

            # ── LLM verdicts ──
            if decision in ("LLM_URLC01", "LLM_MAJ_URLC01"):
                # URLC01 only on company's own domain; third-party CDN → URLC02
                if _v4_is_company_owned(row):
                    return "URLC01"
                return "URLC02"
            if decision in ("LLM_URLC02", "LLM_MAJ_URLC02"):
                return "URLC02"

            # ── Unresolved: company-owned document on own domain → URLC01 ──
            if _v4_is_company_owned(row) and _v4_is_company_document(row):
                return "URLC01"
            if code == "URLC01" and _v4_is_company_owned(row):
                return "URLC01"
            if code == "URLC02":
                return "URLC02"
            return "Manual_Review"

        result["verdict"] = result.apply(_v4_verdict, axis=1)

    elif verdict_mode == "v5":
        # ── V5: Like v4, but allowlisted third-party portals → URLC03 ──
        def _v5_verdict(row):
            decision = row.get("auto_decision", "")
            code = row.get("predicted_url_code", "")

            if decision in ("AUTO_URLC02_GENERIC", "AUTO_URLC02_REGISTRY", "AUTO_URLC02_GENERIC_HOST"):
                if _v5_is_allowlisted_third_party(row):
                    return "URLC03"
                return "URLC02"

            if decision == "AUTO_URLC01_REPORT_HOST":
                return "URLC03"

            if decision == "MANUAL_REVIEW_CHILD_TABLE":
                return "Manual_Review"

            if decision == "MANUAL_REVIEW_SUBSIDIARY":
                return "Manual_Review"

            if decision == "AUTO_URLC01_SUBSIDIARY":
                return "Manual_Review"

            if decision == "MANUAL_REVIEW_UPSTREAM_SUGGESTION":
                return "Manual_Review"

            if decision in (
                "AUTO_URLC01_PARENT_URL",
                "AUTO_URLC01_COMPANY_URL",
                "AUTO_URLC01_MAJORITY_DOMAIN",
            ):
                return "URLC01"

            if decision in ("LLM_URLC01", "LLM_MAJ_URLC01"):
                if _v4_is_company_owned(row):
                    return "URLC01"
                if _v5_is_allowlisted_third_party(row):
                    return "URLC03"
                return "URLC02"
            if decision in ("LLM_URLC02", "LLM_MAJ_URLC02"):
                if _v5_is_allowlisted_third_party(row):
                    return "URLC03"
                return "URLC02"

            if _v4_is_company_owned(row) and _v4_is_company_document(row):
                return "URLC01"
            if code == "URLC01" and _v4_is_company_owned(row):
                return "URLC01"
            if code == "URLC02":
                if _v5_is_allowlisted_third_party(row):
                    return "URLC03"
                return "URLC02"
            return "Manual_Review"

        result["verdict"] = result.apply(_v5_verdict, axis=1)

    else:
        # ── V1 (default): URLC01 & Manual Review only ──
        result["verdict"] = result["predicted_url_code"].replace({
            "UNRESOLVED": "Manual Review",
            "MANUAL_REVIEW": "Manual Review",
            "URLC02": "Manual Review",
        })

    result = add_explanations(result, llm_reasons=llm_reasons, verdict_mode=verdict_mode)

    # ── V2 post-processing: reclassify LLM Manual Review rows ────────────────
    # LLM-detected rows start as Manual Review. If the LLM explanation indicates
    # company-ownership or financial filing → promote to URLC01.
    # Only "Company-specific subdomain/page (third-party-name)" and truly
    # unresolved rows remain Manual Review.
    if verdict_mode == "v2":
        mr_mask = result["verdict"] == "Manual Review"
        if mr_mask.any():
            explanation = result.loc[mr_mask, "explanation"].fillna("")

            # Patterns that indicate company-owned or financial filing → URLC01
            _company_owned_patterns = [
                r"\(company-owned",              # (company-owned CDN), (company-owned hosting), etc.
                r"own domain",                   # "X own domain"
                r"own info site",                # "X own info site"
                r"parent domain",                # parent company domain
                r"parent company domain",
                r"parent \(",                    # "KBFG - Kookmin Bank parent (KB Financial Group)"
                r"subsidiary",                   # subsidiary brand/domain
                r"Domain stem matches issuer name",
                r"Majority domain verified by LLM",
                r"likely company-owned",         # "Short subdomain pattern ... (likely company-owned)"
                r"subdomain$",                   # "PSE - Philip Morris CR subdomain"
            ]

            owned_mask = explanation.str.contains(
                "|".join(_company_owned_patterns), case=False, regex=True
            )

            # Exclude obvious third-party news sites that slip through
            _third_party_override = [
                r"news page$",                   # "KDH News - Dongjin Semichem news page"
            ]
            tp_mask = explanation.str.contains(
                "|".join(_third_party_override), case=False, regex=True
            )

            promote_mask = owned_mask & ~tp_mask
            n_promoted = promote_mask.sum()

            if n_promoted > 0:
                result.loc[mr_mask & promote_mask.reindex(result.index, fill_value=False), "verdict"] = "URLC01"
                log.info(
                    f"V2 post-processing: promoted {n_promoted} LLM Manual Review → URLC01 "
                    f"(company-owned / financial filing). "
                    f"Remaining Manual Review: {(result['verdict'] == 'Manual Review').sum()}"
                )

    # ── V4/V5 post-processing: LLM subsidiary reclassification + Tag column ───
    if verdict_mode in ("v4", "v5"):
        mode_label = verdict_mode.upper()
        # LLM-detected rows that mention "subsidiary" → Manual_Review (not URLC01)
        urlc01_mask = result["verdict"] == "URLC01"
        if urlc01_mask.any():
            explanation = result.loc[urlc01_mask, "explanation"].fillna("")
            _subsidiary_patterns = [
                r"subsidiary",
                r"parent \(",
                r"parent domain",
                r"parent company",
            ]
            sub_mask = explanation.str.contains(
                "|".join(_subsidiary_patterns), case=False, regex=True
            )
            n_demoted = sub_mask.sum()
            if n_demoted > 0:
                result.loc[urlc01_mask & sub_mask.reindex(result.index, fill_value=False), "verdict"] = "Manual_Review"
                log.info(
                    f"{mode_label} post-processing: demoted {n_demoted} URLC01 → Manual_Review "
                    f"(subsidiary relationship detected in explanation)"
                )

        # Upstream DOMAIN_MISMATCH → demote when URL is NOT on company domain
        urlc01_mask = result["verdict"] == "URLC01"
        if urlc01_mask.any():
            mismatch_mask = result.apply(_has_upstream_domain_mismatch, axis=1)
            not_owned = ~result.apply(_v4_is_company_owned, axis=1)
            to_demote = urlc01_mask & mismatch_mask & not_owned
            if verdict_mode == "v5":
                allowlisted = result.apply(_v5_is_allowlisted_third_party, axis=1)
                to_urlc03 = to_demote & allowlisted
                to_urlc02 = to_demote & ~allowlisted
                n_urlc03 = to_urlc03.sum()
                n_urlc02 = to_urlc02.sum()
                if n_urlc03 > 0:
                    result.loc[to_urlc03, "verdict"] = "URLC03"
                    result.loc[to_urlc03, "explanation"] = (
                        "Allowed third-party (URLC03): upstream DOMAIN_MISMATCH on allowlisted portal"
                    )
                if n_urlc02 > 0:
                    result.loc[to_urlc02, "verdict"] = "URLC02"
                    result.loc[to_urlc02, "explanation"] = (
                        "Third-party: upstream DOMAIN_MISMATCH — URL domain is not the issuer's own domain"
                    )
                if n_urlc03 or n_urlc02:
                    log.info(
                        f"{mode_label} post-processing: demoted DOMAIN_MISMATCH "
                        f"{n_urlc03} → URLC03, {n_urlc02} → URLC02"
                    )
            else:
                n_tp = to_demote.sum()
                if n_tp > 0:
                    result.loc[to_demote, "verdict"] = "URLC02"
                    result.loc[to_demote, "explanation"] = (
                        "Third-party: upstream DOMAIN_MISMATCH — URL domain is not the issuer's own domain"
                    )
                    log.info(
                        f"{mode_label} post-processing: demoted {n_tp} URLC01 → URLC02 "
                        f"(upstream DOMAIN_MISMATCH flag)"
                    )

        result = add_v4_tags(result, child_df)

    # ── Merge back into original file ─────────────────────────────────────────
    # Columns to merge: verdict, explanation, plus the 4 suggestion columns
    _suggestion_cols = [
        'URL_CONSISTENCY_SUGGESTED_ISSUER_ID', 'URL_CONSISTENCY_SUGGESTED_ISSUER_NAME',
        'URL_CONSISTENCY_SUGGESTION_URL', 'URL_CONSISTENCY_SUGGESTION_BASIS',
    ]
    merge_cols = ["verdict", "explanation"]
    # Include tag column for v4/v5
    if "tag" in result.columns:
        merge_cols.append("tag")
    # Only include suggestion columns if they exist in result and have any data
    for sc in _suggestion_cols:
        if sc in result.columns:
            merge_cols.append(sc)

    verdict_cols = result[merge_cols].copy()
    verdict_cols.index = fail_df.index

    if pass_df is not None:
        full_df = pd.concat([fail_df, pass_df]).sort_index()
        for col in merge_cols:
            full_df[col] = None
            if not fail_df.empty:
                full_df.loc[fail_df.index, col] = verdict_cols[col].values
        skip_idx = pass_df.index
        full_df.loc[skip_idx, "verdict"] = SKIP_VERDICT
        full_df.loc[skip_idx, "explanation"] = skip_explanation_for_df(pass_df)
        out = full_df
        log.info(
            f"Merged verdicts back into full dataset ({len(out)} rows); "
            f"{len(pass_df)} marked '{SKIP_VERDICT}'"
        )
    else:
        fail_df = fail_df.copy()
        for col in merge_cols:
            # Don't overwrite existing suggestion columns if they already have data from input
            if col in _suggestion_cols and col in fail_df.columns:
                # Only fill where our pipeline set a value (child table / upstream)
                mask = verdict_cols[col].notna()
                fail_df.loc[mask.values, col] = verdict_cols.loc[mask, col].values
            else:
                fail_df[col] = verdict_cols[col].values
        out = fail_df
        log.info(f"Added verdict columns to {len(out)} rows")

    # ── Save output ──────────────────────────────────────────────────────────
    output_path = output_path or str(OUTPUT_DIR / "results.xlsx")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out.to_excel(str(output_path), index=False)
    log.info(f"Results saved to {output_path}")

    format_excel(str(output_path))

    # ── Summary ──────────────────────────────────────────────────────────────
    classified = out[out["verdict"].notna()]
    vc = classified["verdict"].value_counts()
    log.info(f"\n{'='*40}\n  SUMMARY\n{'='*40}")
    log.info(f"  Total rows in file: {len(out)}")
    log.info(f"  Rows classified: {len(classified)}")
    for v, c in vc.items():
        log.info(f"    {v}: {c}")
    if pass_df is not None:
        log.info(f"  Skipped rows ({SKIP_VERDICT}): {len(pass_df)}")

    summary = {
        "total_rows": len(out),
        "classified_rows": len(classified),
        "verdicts": vc.to_dict(),
        "passing_rows": len(pass_df) if pass_df is not None else 0,
    }

    # ── Evaluation ───────────────────────────────────────────────────────────
    if run_eval and "COMMENT_CODE" in result.columns:
        metrics = evaluate_against_ground_truth(result)
        log.info(f"\n{'='*40}\n  EVALUATION vs GROUND TRUTH\n{'='*40}")
        log.info(f"  Accuracy: {metrics['accuracy']:.1f}%")
        log.info(f"  Resolved: {metrics['resolved_rows']}/{metrics['total_rows']} "
                 f"({metrics['resolution_rate']:.1f}%)")
        summary["eval"] = metrics

    return summary
